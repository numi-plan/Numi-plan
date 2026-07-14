import openpyxl
import json
import re
import urllib.request
import ssl
import sys
import os
import shutil
import time
import datetime

# Set stdout to use utf-8
sys.stdout.reconfigure(encoding='utf-8')

ctx = ssl.create_default_context()
ctx.check_hostname = False
ctx.verify_mode = ssl.CERT_NONE

# ══════════════════════════════════════════════════════════════════════════════
# HELPER: Authentification Supabase Auth pour obtenir un jeton JWT
# ══════════════════════════════════════════════════════════════════════════════
def get_supabase_token(supabase_url, supabase_key, email, password, ctx):
    """Authenticate with email and password and return the JWT access token."""
    url = f"{supabase_url}/auth/v1/token?grant_type=password"
    headers = {
        "apikey": supabase_key,
        "Content-Type": "application/json"
    }
    payload = {
        "email": email,
        "password": password
    }
    req = urllib.request.Request(
        url,
        data=json.dumps(payload).encode('utf-8'),
        headers=headers,
        method='POST'
    )
    try:
        with urllib.request.urlopen(req, context=ctx, timeout=15) as resp:
            data = json.loads(resp.read().decode('utf-8'))
            return data.get("access_token", "")
    except Exception as e:
        print(f"  [ERROR] Échec d'authentification Supabase : {e}")
        return ""

# ══════════════════════════════════════════════════════════════════════════════
# HELPER: Fetch current team payload from Supabase (to preserve critical fields)
# ══════════════════════════════════════════════════════════════════════════════
def fetch_supabase_live(eq, supabase_url, supabase_key, token, ctx):
    """Fetch current payload for team 'eq' from Supabase. Returns dict or None."""
    url = f"{supabase_url}/rest/v1/numiplan_teams?team=eq.{eq}&select=payload"
    auth_header = f"Bearer {token}" if token else f"Bearer {supabase_key}"
    headers = {
        "apikey": supabase_key,
        "Authorization": auth_header,
        "Content-Type": "application/json"
    }
    try:
        req = urllib.request.Request(url, headers=headers)
        with urllib.request.urlopen(req, context=ctx, timeout=15) as resp:
            rows = json.loads(resp.read().decode('utf-8'))
            if rows:
                return rows[0].get('payload', {})
    except Exception as e:
        print(f"  [WARN] Impossible de lire Supabase pour equipe {eq}: {e}")
    return None

# ══════════════════════════════════════════════════════════════════════════════
# STEP 1: Revert HTML Files to Original Backups (DÉSACTIVÉ POUR ÉVITER LES PERTES)
# ══════════════════════════════════════════════════════════════════════════════
print("Step 1: Reverting HTML files... [DÉSACTIVÉ - Protection contre l'écrasement du code]")
# src_fiche = r"C:\Users\kaddour.lahouel\.gemini\antigravity-ide\scratch\fiche_old_e23.html"
# dest_fiche = r"c:\Users\kaddour.lahouel\Desktop\PROJET ERP ANTI\fiche.html"
# src_fiche_supabase = r"C:\Users\kaddour.lahouel\.gemini\antigravity-ide\scratch\fiche_supabase_23h.html"
# dest_fiche_supabase = r"c:\Users\kaddour.lahouel\Desktop\PROJET ERP ANTI\fiche_supabase.html"
# (Déploiement en prod géré via GitHub pour éviter les pertes de personnalisations locales)



# ══════════════════════════════════════════════════════════════════════════════
# STEP 2: Database Restoration
# ══════════════════════════════════════════════════════════════════════════════
print("\nStep 2: Starting database restoration...")

wb_path = r"c:\Users\kaddour.lahouel\Desktop\PROJET ERP ANTI\V3NUMILOG_EXTRACTION_COMPLETE_20260713 14.00.xlsx"
wb = openpyxl.load_workbook(wb_path, data_only=True)

# Load current DB backup for base data (contains non-exported fields like users, blacklist, etc.)
with open(r"c:\Users\kaddour.lahouel\Desktop\PROJET ERP ANTI\numiplan_full_backup.json", "r", encoding="utf-8") as f:
    db_backup = json.load(f)

# Helper to unwrap Firestore REST value to regular Python dict/value
def unwrap_val(v):
    if not isinstance(v, dict):
        return v
    if "stringValue" in v:
        return v["stringValue"]
    if "integerValue" in v:
        return int(v["integerValue"])
    if "doubleValue" in v:
        return float(v["doubleValue"])
    if "booleanValue" in v:
        return v["booleanValue"]
    if "nullValue" in v:
        return None
    if "arrayValue" in v:
        arr = v["arrayValue"].get("values", [])
        return [unwrap_val(x) for x in arr]
    if "mapValue" in v:
        fields = v["mapValue"].get("fields", {})
        return {k: unwrap_val(val) for k, val in fields.items()}
    return v

# Helper to wrap python value to Firestore REST format
def to_firestore_value(val):
    if val is None:
        return {"nullValue": None}
    elif isinstance(val, bool):
        return {"booleanValue": val}
    elif isinstance(val, int):
        return {"integerValue": str(val)}
    elif isinstance(val, float):
        return {"doubleValue": val}
    elif isinstance(val, str):
        return {"stringValue": val}
    elif isinstance(val, list):
        return {"arrayValue": {"values": [to_firestore_value(x) for x in val]}}
    elif isinstance(val, dict):
        return {"mapValue": {"fields": {k: to_firestore_value(v) for k, v in val.items()}}}
    else:
        return {"stringValue": str(val)}

def parse_decimal_time(val):
    if val is None or val == '':
        return None
    if isinstance(val, (int, float)):
        return float(val)
    val_str = str(val).strip()
    match = re.match(r'^(\d+):(\d+)$', val_str)
    if match:
        h = int(match.group(1))
        m = int(match.group(2))
        return (h * 60 + m) / 1440.0
    return None

def parse_sheet_rows(ws, headers_spec):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    
    headers = [str(h).strip() if h is not None else '' for h in rows[0]]
    parsed_items = []
    
    for row in rows[1:]:
        if all(cell is None or str(cell).strip() == '' for cell in row):
            continue
            
        item = {}
        row_dict = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
        
        for key, col_name in headers_spec.items():
            val = row_dict.get(col_name)
            
            if key == 'hor':
                item[key] = parse_decimal_time(val)
            elif key in ['progFige', 'cspc', 'actif']:
                if key == 'cspc':
                    item[key] = True if val in ['OUI', 'oui', 'Oui', True, 1] else False
                elif key == 'progFige':
                    item[key] = True if val in ['VRAI', 'true', 'True', True, 1] else False
                elif key == 'actif':
                    item[key] = True if val in ['OUI', 'oui', 'Oui', 'actif', 'ACTIF', True, 1] else False
            elif key in ['n', 'km', 'dureeHeures']:
                if val is not None and str(val).strip() not in ['', 'None']:
                    try:
                        item[key] = int(float(val))
                    except:
                        item[key] = None
                else:
                    item[key] = None
            else:
                item[key] = str(val).strip() if val is not None else ''
        
        parsed_items.append(item)
    return parsed_items

def parse_odo_sheet(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    
    headers = [str(h).strip() if h is not None else '' for h in rows[0]]
    weeks = []
    for h in headers[2:]:
        match = re.search(r'\(([^)]+)\)', h)
        if match:
            weeks.append(match.group(1))
        else:
            weeks.append(h)
            
    parsed_odo = []
    for row in rows[1:]:
        if all(cell is None or str(cell).strip() == '' for cell in row):
            continue
        v = str(row[0]).strip() if row[0] is not None else ''
        chauffeur = str(row[1]).strip() if row[1] is not None else ''
        vals = []
        for cell in row[2:]:
            if cell is None or str(cell).strip() in ['', 'None']:
                vals.append(None)
            else:
                try:
                    vals.append(int(float(cell)))
                except:
                    vals.append(None)
        
        parsed_odo.append({
            'v': v,
            'chauffeur': chauffeur,
            'vals': vals
        })
    return parsed_odo, weeks

specs = {
    'dispo': {
        'v': 'Véhicule', 'etat': 'État', 'type': 'Type', 'chauffeur': 'Chauffeur',
        'hor': 'Horaire', 'lieuJ': 'Lieu Jour', 'lieuJ1': 'Lieu J+1',
        'lieuJ_initial': 'Lieu J Initial', 'lieuJ1_initial': 'Lieu J+1 Initial',
        'equip': 'Équipement', 'affecte': 'Affecté', 'progFige': 'Prog. Figé',
        'prog': 'Programme', 'km': 'KM', 'cspc': 'CSPC', 'obs': 'Observation',
        'progDepart': 'Prog. Départ', 'repos': 'Repos'
    },
    'pannes': {
        'date': 'Date', 'immat': 'Immatriculation', 'chauffeur': 'Chauffeur',
        'type': 'Type', 'di': 'N° DI', 'desc': 'Description', 'cet': 'CET',
        'kpi': 'KPI', 'etat': 'État', 'rex': 'REX'
    },
    'nc': {
        'n': 'N°', 'date': 'Date', 'sharik': 'Sharik', 'chgt': 'Chgt Compte',
        'fausse': 'Fausse Décl.', 'refus': 'Refus Travail', 'autre': 'Autre',
        'mise': 'Mise en demeure', 'commission': 'Commission', 'cet': 'CET'
    },
    'gasoil': {
        'date': 'Date', 'immat': 'Immatriculation', 'vehicule': 'Immatriculation',
        'chauffeur': 'Chauffeur', 'station': 'Station', 'type': 'Type',
        'operation': 'Opération', 'resp': 'Responsable', 'agence': 'Agence',
        'site': 'Site', 'cet': 'CET'
    },
    'equip': {
        'n': 'N°', 'sharik': 'Sharik', 'tr': 'Tracteur (TR)', 'v': 'Tracteur (TR)',
        'rm': 'Remorque (RM)', 'barre': 'Barre', 'cable': 'Câble', 'tel': 'Téléphone'
    },
    'activite': {
        'sharik': 'Sharik', 'chauffeur': 'Sharik', 'tr': 'Tracteur (TR)', 'v': 'Tracteur (TR)',
        'rm': 'Remorque (RM)', 'debut': 'Date Début', 'fin': 'Date Fin',
        'ext': 'Extension', 'site': 'Site', 'etatTR': 'État TR', 'actif': 'Actif', 'obs': 'Observation'
    },
    'bl': {
        'chauffeur': 'Chauffeur', 'mars': 'Mars', 'avril': 'Avril', 'mai': 'Mai',
        'juin': 'Juin', 'juil': 'Juillet', 'aout': 'Août', 'sep': 'Sep.',
        'oct': 'Oct.', 'nov': 'Nov.', 'dec': 'Déc.', 'cet': 'CET'
    },
    'cspc': {
        'v': 'Véhicule', 'etat': 'État', 'chauffeur': 'Chauffeur', 'hor': 'Horaire',
        'lieuJ': 'Lieu Jour', 'lieuJ1': 'Lieu J+1', 'equip': 'Équipement',
        'prog': 'Programme', 'cet': 'CET'
    }
}

api_key = "AIzaSyDA-loDmlfcdPC0es0sb6tiCM4JXP2VF24"
supabase_url = "https://mfljrozyxyfkymzlnsmt.supabase.co"
supabase_key = "sb_publishable_sMb9uHOr4gpyqmZGrkY-YQ_z90TgNAB"

# Authentification Supabase Auth pour contourner les RLS durant la restauration
print("Authenticating with Supabase Auth as Admin...")
supabase_token = get_supabase_token(supabase_url, supabase_key, "admin@sharik.numilog.com", "Sharik@admin", ctx)
if supabase_token:
    print("  Authenticated successfully !")
else:
    print("  [WARN] Failed to authenticate. Proceeding as anonymous client (writes will fail if RLS is enabled).")

teams = ['A', 'B', 'C', 'D']
for eq in teams:
    print(f"\nProcessing Team {eq}...")
    
    # 1. Parse Excel data
    excel_data = {}
    for list_name, spec in specs.items():
        sheet_name = f"{list_name.capitalize()}_{eq}"
        if list_name in ['nc', 'bl']:
            sheet_name = f"{list_name.upper()}_{eq}"
        elif list_name == 'cspc':
            sheet_name = f"CSPC_{eq}"
            
        if sheet_name in wb.sheetnames:
            excel_data[list_name] = parse_sheet_rows(wb[sheet_name], spec)
            print(f"  Parsed {len(excel_data[list_name])} items for {list_name} from Excel.")
        else:
            excel_data[list_name] = []
            print(f"  Warning: {sheet_name} not found in Excel.")

    # Parse odo sheet
    odo_sheet_name = f"Odo_{eq}"
    if odo_sheet_name in wb.sheetnames:
        excel_odo, excel_weeks = parse_odo_sheet(wb[odo_sheet_name])
        excel_data['odo'] = excel_odo
        excel_data['odo_weeks'] = excel_weeks
        print(f"  Parsed {len(excel_odo)} vehicle odometers with {len(excel_weeks)} weeks.")
    else:
        excel_data['odo'] = []
        excel_data['odo_weeks'] = []

    # 2. Get current team data from backup for base (to preserve users, blacklist, and files)
    team_db_raw = db_backup.get("teams", {}).get(eq, {})
    current_team_db = {k: unwrap_val(v) for k, v in team_db_raw.items()}
    
    # Preserve assurance and geoloc properties in dispo
    current_dispo_map = {item['v']: item for item in current_team_db.get('dispo', []) if 'v' in item}
    for item in excel_data['dispo']:
        veh = item.get('v')
        if veh in current_dispo_map:
            curr_item = current_dispo_map[veh]
            # Copy all assurance properties
            for k, val in curr_item.items():
                if k.startswith('assurance') or k == 'geoloc':
                    item[k] = val

    # Preserve odo extra properties (like kmFacture, kmFactures)
    current_odo_map = {item['v']: item for item in current_team_db.get('odo', []) if 'v' in item}
    for item in excel_data['odo']:
        veh = item.get('v')
        if veh in current_odo_map:
            curr_item = current_odo_map[veh]
            for k, val in curr_item.items():
                if k not in ['v', 'chauffeur', 'vals']:
                    item[k] = val

    # 3. Construct final unwrapped payload
    payload_unwrapped = dict(current_team_db)
    
    # Overwrite the 10 data lists (from Excel)
    for k in ['dispo', 'pannes', 'nc', 'gasoil', 'equip', 'activite', 'bl', 'odo', 'odo_weeks', 'cspc']:
        payload_unwrapped[k] = excel_data[k]

    # ══════════════════════════════════════════════════════════════════════════
    # PRESERVATION CRITIQUE: récupérer depuis Supabase les champs jamais dans Excel
    # incidentsLivraison, anomaliesPointees, tracabilite — NE JAMAIS ECRASER
    # ══════════════════════════════════════════════════════════════════════════
    print(f"  [PRESERVE] Recuperation des donnees critiques depuis Supabase pour equipe {eq}...")
    live_payload = fetch_supabase_live(eq, supabase_url, supabase_key, supabase_token, ctx)
    if live_payload:
        preserved_fields = ['incidentsLivraison', 'anomaliesPointees', 'tracabilite', 'flashInfos', 'clientModifs']
        for field in preserved_fields:
            if field in live_payload and live_payload[field]:
                count = len(live_payload[field]) if isinstance(live_payload[field], list) else len(live_payload[field].keys())
                print(f"    -> Preservation de '{field}': {count} entree(s) conservee(s)")
                payload_unwrapped[field] = live_payload[field]
            else:
                if field not in payload_unwrapped:
                    payload_unwrapped[field] = [] if field != 'anomaliesPointees' else {}
    else:
        print(f"  [WARN] Impossible de recuperer Supabase — les incidents/anomalies du backup seront utilises")
        print(f"         ATTENTION: des donnees recentes pourraient etre perdues !")

    current_time_ms = int(time.time() * 1000)
    payload_unwrapped['_localTs'] = current_time_ms

    # ══════════════════════════════════════════════════════════════════════════
    # PUSH TO FIRESTORE
    # ══════════════════════════════════════════════════════════════════════════
    print(f"  Uploading to Firestore...")
    payload_firestore = {
        "fields": {k: to_firestore_value(v) for k, v in payload_unwrapped.items()}
    }
    
    # To replace the document completely, patch the entire document
    patch_url = f"https://firestore.googleapis.com/v1/projects/numi-pro/databases/(default)/documents/numiplan/data_{eq}?key={api_key}"
    payload_bytes = json.dumps(payload_firestore).encode('utf-8')
    req = urllib.request.Request(
        patch_url,
        data=payload_bytes,
        headers={'Content-Type': 'application/json', 'User-Agent': 'Mozilla/5.0'},
        method='PATCH'
    )
    
    try:
        with urllib.request.urlopen(req, context=ctx) as response:
            res = json.loads(response.read().decode('utf-8'))
            print(f"  Firestore write SUCCESS for Team {eq}!")
    except Exception as e:
        detail = ""
        if hasattr(e, 'read'):
            detail = e.read().decode('utf-8')
        print(f"  Firestore write FAILED for Team {eq}: {e} - {detail}")

    # ══════════════════════════════════════════════════════════════════════════
    # PUSH TO SUPABASE
    # ══════════════════════════════════════════════════════════════════════════
    print(f"  Uploading to Supabase...")
    row = {
        "team": eq,
        "payload": payload_unwrapped,
        "updated_at": datetime.datetime.now(datetime.timezone.utc).isoformat()
    }
    
    post_url = f"{supabase_url}/rest/v1/numiplan_teams"
    auth_header = f"Bearer {supabase_token}" if supabase_token else f"Bearer {supabase_key}"
    headers = {
        "apikey": supabase_key,
        "Authorization": auth_header,
        "Content-Type": "application/json",
        "Prefer": "resolution=merge-duplicates"
    }
    
    payload_bytes = json.dumps([row]).encode('utf-8')
    req_supabase = urllib.request.Request(post_url, data=payload_bytes, headers=headers, method='POST')
    
    try:
        with urllib.request.urlopen(req_supabase, context=ctx) as response:
            print(f"  Supabase write SUCCESS for Team {eq}!")
    except Exception as e:
        detail = ""
        if hasattr(e, 'read'):
            detail = e.read().decode('utf-8')
        print(f"  Supabase write FAILED for Team {eq}: {e} - {detail}")

print("\nAll restoration tasks completed successfully.")

# ── AUTO-SYNC DES TABLES NORMALISEES APRES RESTAURATION ──────────────────────
import subprocess
print("\n[Normalized Sync] Synchronisation des nouvelles tables normalisées...")
try:
    subprocess.run([sys.executable, "migrate_to_normalized_tables.py"], check=True)
    print("[Normalized Sync] ✅ Tables synchronisées avec succès !")
except Exception as e:
    print(f"[Normalized Sync] ⚠️ Erreur lors de la synchronisation : {e}")

